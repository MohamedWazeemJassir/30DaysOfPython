import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

sheet["A1"] = "Products"
sheet["B1"] = "Price"

sheet.cell(row=2, column=1, value="Laptop")
sheet.cell(row=2, column=2, value=45000)

wb.save("new_report.xlsx")
